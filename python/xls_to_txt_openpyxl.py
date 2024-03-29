#!/usr/bin/python3

#    Excel to text files export script. Exports all tabs with a valid header.
#    2020 Victor Barashko, University of Florida/Physics
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <https://www.gnu.org/licenses/>.


import openpyxl as pyxl
import os
import sys
from pathlib import Path

# XLSX config file
loc = ("instantiations_tabfiles.xlsx")

nargs = len(sys.argv)

if nargs > 1:
    loc = sys.argv[1] 


wb = pyxl.load_workbook(loc,data_only=True)
#print("Number of sheets in ",loc, ": ", wb.nsheets)

for sheet in wb:
    print("Sheet: ", sheet.title)
    # Check for header fields 
    path = ""
    fn = ""
    ncols = 0    
    
    if ("// path" in str(sheet.cell(row=1,column=1).value)):
        path = sheet.cell(row=1,column=2).value
        print(sheet.cell(row=1,column=1).value, path)
    if ("// file" in str(sheet.cell(row=2,column=1).value)):
        fn = sheet.cell(row=2,column=2).value
        print(sheet.cell(row=2,column=1).value, fn)
    if ("// columns" in str(sheet.cell(row=3,column=1).value)):
        ncols = sheet.cell(row=3,column=2).value
        print(sheet.cell(row=3,column=1).value, ncols)
    # Header is found. Proceed with writing text file
    if ((path != "") and (fn != "") and (ncols != 0)):
        
        out_dir = Path(path)        
        out_txt = out_dir / fn
        
        print("Converting to txt file", out_txt, "with", ncols, "columns")
        
        if not os.path.exists(out_dir):
            os.makedirs(out_dir)
            
        with open(out_txt, "w") as f:
            cnt = 0
            for i_row in range(4,sheet.max_row+1):
                # Skip comment lines with "//"
                #if ("//" in str(sheet.cell(row=i_row,column=1).value)):
                #    continue
                    
                out_row = ""
                for col in range(1,ncols+1):                    
                    val = str(sheet.cell(row=i_row,column=col).value)
                    if (sheet.cell(row=i_row,column=col).value is None):
                        val = ""
                    out_row = out_row + val + "\t"
                cnt = cnt + 1 
                out_row = out_row + "\n"
                # print(out_row)
                f.write(out_row)
                
            print("Wrote ", cnt, "lines")
            f.close()

    else:
        print("No header found. Skipped conversion")
    print("--------------------")
    
print("Done")
